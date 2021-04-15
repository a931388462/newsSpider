#!/usr/bin/python
# -*- coding: UTF-8 -*-
import datetime
import time
from tkinter import *  # 导入 Tkinter 库
import tkinter as tk
from tkinter import ttk, messagebox

from openpyxl import Workbook

from baiduTransAPI import translate, transText
from myCalendar import Calendar
from news import News
from spider import spider1, getNewstext1, spider2, getNewstext2, spider3, getNewstext3, getNewstext4, spider4, spider5, \
    getNewstext5

root = Tk()
# 设置窗口的大小宽x高+偏移量
#root.geometry('1200x1600')
# 设置窗口标题
root.title('爬虫程序')

  #创建frame容器
frmLT = Frame(width=1200, height=50)
frmLM = Frame(width=1200)
frmLB = Frame(width=1200, height=50)

frmLT.grid(row=0, column=0,padx=1,pady=3)
frmLM.grid(row=1, column=0,padx=1,pady=3)
frmLB.grid(row=2, column=0,padx=1,pady=10)

#网址
Label(frmLT,text='网站').grid(row=0,sticky=W)
# 创建下拉菜单
cmb = ttk.Combobox(frmLT)
cmb.grid(row=0,column=1)
# 设置下拉菜单中的值
cmb['value'] = ('1','2','3','4','5')
# 设置默认值，即默认下拉框中的内容
cmb.current(0)
# 默认值中的内容为索引，从0开始
# 执行函数
def func(event):
    print(cmb.get())
cmb.bind("<<ComboboxSelected>>",func)

#Calendar位置
width, height = frmLT.winfo_reqwidth() + 50, 50  # 窗口大小
x, y = (frmLT.winfo_screenwidth() - width) / 2, (frmLT.winfo_screenheight() - height) / 2

#开始日期
date_str_start = tk.StringVar()
#初始值
date_str_start.set((datetime.datetime.now()+datetime.timedelta(days=-1)).strftime("%Y-%m-%d"))
start_date_text = ttk.Entry(frmLT, textvariable=date_str_start, state='disabled')
start_date_text.grid(row=0,column=3)
# Calendar((x, y), 'ur').selection() 获取日期，x,y为点坐标
date_str_gain = lambda: [
    date_str_start.set(date)
    for date in [Calendar((x, y), 'ur').selection()]
    if date]
tk.Button(frmLT, text='开始日期:', command=date_str_gain).grid(row=0, column=2)

#结束日期
date_str_end = tk.StringVar()
#初始值
date_str_end.set((datetime.datetime.now()).strftime("%Y-%m-%d"))
end_date_text = ttk.Entry(frmLT, textvariable=date_str_end, state='disabled')
end_date_text.grid(row=0,column=5)
# Calendar((x, y), 'ur').selection() 获取日期，x,y为点坐标
date_end_gain = lambda: [
    date_str_end.set(date)
    for date in [Calendar((x, y), 'ur').selection()]
    if date]
tk.Button(frmLT, text='结束日期:', command=date_end_gain).grid(row=0, column=4)


#表格
tree=ttk.Treeview(frmLM, show = "headings", selectmode = tk.BROWSE,height=25)
tree.grid()
#定义列
tree["columns"]=("date","website","title","abstract","text")
#设置列属性，列不显示
tree.column("date",width=100)
tree.column("website",width=100)
tree.column("title",width=100)
tree.column("abstract",width=650)
tree.column("text",width=200)

#设置表头
tree.heading("date",text="日期（date）")
tree.heading("website",text="网站（website）")
tree.heading("title",text="标题（title）")
tree.heading("abstract",text="摘要（abstract）")
tree.heading("text",text="正文（text）")

#点击搜索调用
def execute():
    web_type = cmb.get()
    newss = []
    print(cmb.get())
    if web_type == '1':
        newss = spider1(date_str_start.get(),date_str_end.get())
    elif web_type == '2':
        newss = spider2(date_str_start.get(), date_str_end.get())
    elif web_type == '3':
        newss = spider3(date_str_start.get(), date_str_end.get())
    elif web_type == '4':
        newss = spider4(date_str_start.get(), date_str_end.get())
    elif web_type == '5':
        newss = spider5(date_str_start.get(), date_str_end.get())
    outResult(newss,tree)

#搜索
button_search = Button(frmLT, text='搜索', width = 8 ,command=execute)
button_search.grid(row=0,column=7)

#向表格输出
def outResult(newss,tree):
    #清空已有数据
    x = tree.get_children()
    for item in x:
        tree.delete(item)
    i = 0
    for news in newss:
        #添加数据  0 为列的下标，第0行
        tree.insert("",0,values=(news.date,news.url,translate(news.title),translate(news.summary),"双击查看正文"))
        i+=1

#正文调度器
def textScheduler(url):
        web_type = cmb.get()
        newstext = ''
        if web_type == '1':
            newstext = getNewstext1(url)
        elif web_type == '2':
            newstext = getNewstext2(url)
        elif web_type == '3':
            newstext = getNewstext3(url)
        elif web_type == '4':
            newstext = getNewstext4(url)
        elif web_type == '5':
            newstext = getNewstext5(url)
        print(newstext)
        return newstext

# 获取当前点# 击行的值
def treeviewClick(event):  # 单击
    for item in tree.selection():
        item_text = tree.item(item, "values")
        # 弹出对话框
    messagebox.showinfo(title = '新闻内容',message = transText(textScheduler(item_text[1])))
# 鼠标左键双击
tree.bind('<Double-Button-1>', treeviewClick)

def outToExcel():
    dateStr = time.strftime("%Y%m%d", time.localtime())
    fileName = date_str_start.get() + "到" + date_str_end.get() + "的新闻_" + dateStr + ".xlsx"
    # 创建文件对象
    wb = Workbook()
    # 获取第一个sheet
    ws = wb.active
    # 表头
    ws['A1'] = "日期"
    ws['B1'] = "网址"
    ws['C1'] = "标题"
    ws['D1'] = "摘要"
    ws['E1'] = "正文"
    # 调整列宽
    #ws.column_dimensions['A'].width = 15.0
    # 获取第一个sheet
    ws = wb.active
    # 写入多个单元格
    items = tree.get_children()
    for item in items:
        item_text = tree.item(item, "values")
        ws.append([item_text[0],item_text[1],item_text[2],item_text[3],transText(textScheduler(item_text[1]))])
        print(item_text)
    # 保存为爬取结果
    wb.save(fileName)
    # 弹出对话框
    messagebox.showinfo(title='导出成功', message="导出成功")

#导出文件
button_search = Button(frmLB, text='导出文件', width = 8, command=outToExcel)
button_search.grid(row=0,column=7)

root.mainloop()